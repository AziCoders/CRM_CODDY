"""Модуль для генерации отчетов по оплатам"""
import json
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional
from datetime import datetime
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from bot.config import ROOT_DIR, CITY_MAPPING, CITIES

# Константы
PRICE_PER_MONTH = 5000
DISCOUNT_PER_CHILD = 500
MIN_CHILDREN_FOR_DISCOUNT = 2

# Статусы оплаты
STATUS_PAID = "Оплатил"
STATUS_WROTE = "Написали"
STATUS_NOT_PAID = "Не оплатил"
STATUS_DEFERRED = "Отсрочка"
STATUS_NOT_STUDYING = "Не учился"

# Поля для исключения из месяцев
EXCLUDED_FIELDS = {"Дата оплаты", "ФИО", "Phone", "Комментарий"}


class PaymentsReportGenerator:
    """Генератор отчетов по оплатам"""

    def __init__(self, city_name: str):
        """
        Инициализация генератора отчетов
        
        Args:
            city_name: Название города (русское)
        """
        self.city_name = city_name
        self.city_en = CITY_MAPPING.get(city_name, city_name)
        self.base_path = ROOT_DIR / f"data/{self.city_en}"

        # Данные
        self.students_data: Dict[str, Any] = {}
        self.payments_data: Dict[str, Any] = {}
        self.groups_data: Dict[str, Any] = {}

        # Объединенные данные
        self.merged_data: List[Dict[str, Any]] = []

        # Месяцы (динамические)
        self.months: List[str] = []

        # Статистика
        self.stats: Dict[str, int] = {
            "paid": 0,
            "wrote": 0,
            "not_paid": 0,
            "deferred": 0
        }

        # Финансовые показатели
        self.finances: Dict[str, int] = {
            "expected_turnover": 0,
            "current_turnover": 0,
            "debt": 0
        }

    def load_data(self) -> None:
        """Загружает данные из JSON файлов"""
        # Загружаем students.json
        students_path = self.base_path / "students.json"
        if students_path.exists():
            with open(students_path, "r", encoding="utf-8") as f:
                self.students_data = json.load(f)

        # Загружаем payments.json
        payments_path = self.base_path / "payments.json"
        if payments_path.exists():
            with open(payments_path, "r", encoding="utf-8") as f:
                self.payments_data = json.load(f)

        # Загружаем groups.json (опционально)
        groups_path = self.base_path / "groups.json"
        if groups_path.exists():
            with open(groups_path, "r", encoding="utf-8") as f:
                self.groups_data = json.load(f)

    def _get_current_month(self) -> str:
        """Возвращает название текущего месяца, который есть в self.months."""
        month_order = [
            "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
            "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
        ]

        now_month_index = datetime.now().month - 1  # 0..11
        current_name = month_order[now_month_index]

        if current_name in self.months:
            return current_name

        # если в данных нет текущего календарного месяца — берём последний из имеющихся
        return self.months[-1] if self.months else current_name

    def _extract_months(self) -> List[str]:
        """Извлекает список всех месяцев из payments.json"""
        months = set()

        if "fields" in self.payments_data:
            for field in self.payments_data["fields"]:
                if field not in EXCLUDED_FIELDS:
                    months.add(field)

        if "payments" in self.payments_data:
            for payment in self.payments_data["payments"]:
                payments_data = payment.get("payments_data", {})
                for month in payments_data.keys():
                    months.add(month)

        # Сортируем месяцы по порядку
        month_order = [
            "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
            "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
        ]

        sorted_months = []
        for month in month_order:
            if month in months:
                sorted_months.append(month)

        # Добавляем месяцы, которых нет в стандартном списке
        for month in sorted(months):
            if month not in sorted_months:
                sorted_months.append(month)

        return sorted_months

    def _find_student_in_students(self, student_id: str, fio: str) -> Optional[Dict[str, Any]]:
        """Находит ученика в students.json по student_id или ФИО"""
        if not self.students_data:
            return None

        # Сначала ищем по student_id
        for group_id, group_data in self.students_data.items():
            students_list = group_data.get("students", [])
            for student in students_list:
                if student.get("ID") == student_id:
                    return student

        # Если не нашли по ID, ищем по ФИО
        fio_clean = fio.strip().lower()
        for group_id, group_data in self.students_data.items():
            students_list = group_data.get("students", [])
            for student in students_list:
                student_fio = student.get("ФИО", "").strip().lower()
                if student_fio == fio_clean:
                    return student

        return None

    def merge_data(self) -> None:
        """Объединяет данные из students.json и payments.json"""
        self.months = self._extract_months()

        # Создаем карту учеников из students.json
        students_map: Dict[str, Dict[str, Any]] = {}
        for group_id, group_data in self.students_data.items():
            students_list = group_data.get("students", [])
            for student in students_list:
                student_id = student.get("ID", "")
                fio = student.get("ФИО", "").strip()
                if student_id:
                    students_map[student_id] = student
                if fio:
                    students_map[fio.lower()] = student

        # Обрабатываем payments
        payments_list = self.payments_data.get("payments", [])

        for payment in payments_list:
            student_id = payment.get("student_id", "")
            fio = payment.get("ФИО", "").strip()
            phone = payment.get("Phone", "").strip()
            comment = payment.get("Комментарий", "").strip()
            payment_url = payment.get("payment_url", "")
            payments_data = payment.get("payments_data", {})

            # Ищем ученика в students.json
            student_info = None
            if student_id:
                student_info = self._find_student_in_students(student_id, fio)
            elif fio:
                student_info = self._find_student_in_students("", fio)

            # Получаем student_url
            student_url = ""
            if student_info:
                student_url = student_info.get("student_url", "")

            # Формируем статусы по месяцам
            month_statuses: Dict[str, str] = {}
            for month in self.months:
                status = payments_data.get(month, "").strip()
                if status:
                    month_statuses[month] = status
                else:
                    month_statuses[month] = STATUS_NOT_STUDYING

            # Добавляем объединенные данные
            merged_record = {
                "fio": fio,
                "student_id": student_id,
                "student_url": student_url,
                "payment_url": payment_url,
                "phone": phone,
                "comment": comment,
                "months": month_statuses
            }

            self.merged_data.append(merged_record)

        # Добавляем учеников из students.json, которых нет в payments.json
        processed_student_ids = {record.get("student_id", "") for record in self.merged_data}
        processed_fios = {record.get("fio", "").lower() for record in self.merged_data}

        for group_id, group_data in self.students_data.items():
            students_list = group_data.get("students", [])
            for student in students_list:
                student_id = student.get("ID", "")
                fio = student.get("ФИО", "").strip()

                # Проверяем, не обработан ли уже
                if student_id and student_id in processed_student_ids:
                    continue
                if fio and fio.lower() in processed_fios:
                    continue

                # Добавляем с статусом "Не учился" для всех месяцев
                month_statuses = {month: STATUS_NOT_STUDYING for month in self.months}

                merged_record = {
                    "fio": fio,
                    "student_id": student_id,
                    "student_url": student.get("student_url", ""),
                    "payment_url": "",
                    "phone": student.get("Номер родителя", "").strip(),
                    "comment": "",
                    "months": month_statuses
                }

                self.merged_data.append(merged_record)

    def calculate_stats(self, month: Optional[str] = None) -> None:
        """Рассчитывает статистику по оплатам за один месяц (по умолчанию — текущий)."""
        if month is None:
            month = self._get_current_month()

        self.stats = {
            "paid": 0,
            "wrote": 0,
            "not_paid": 0,
            "deferred": 0
        }

        for record in self.merged_data:
            months_data = record.get("months", {})
            status = months_data.get(month, STATUS_NOT_STUDYING)

            if status == STATUS_PAID:
                self.stats["paid"] += 1
            elif status == STATUS_WROTE:
                self.stats["wrote"] += 1
            elif status == STATUS_NOT_PAID:
                self.stats["not_paid"] += 1
            elif status == STATUS_DEFERRED:
                self.stats["deferred"] += 1

    def _count_duplicate_children_for_month(self, month: str, only_paid: bool = False) -> int:
        """
        Считает количество детей в семьях с 2+ детьми за один конкретный месяц.
        Если only_paid=True — считаем только тех, кто в этом месяце имеет статус 'Оплатил'.
        """
        phone_counter = Counter()

        for record in self.merged_data:
            months_data = record.get("months", {})
            status = months_data.get(month, STATUS_NOT_STUDYING)

            if status == STATUS_NOT_STUDYING:
                continue  # в этом месяце не учился — в расчёт не идёт

            if only_paid and status != STATUS_PAID:
                continue

            phone = record.get("phone", "").strip()
            if phone:
                phone_counter[phone] += 1

        total_children = 0
        for phone, count in phone_counter.items():
            if count >= MIN_CHILDREN_FOR_DISCOUNT:
                total_children += count

        return total_children

    def _count_paid_students(self) -> int:
        """Подсчитывает количество учеников, которые оплатили хотя бы один месяц"""
        paid_students = set()

        for record in self.merged_data:
            months_data = record.get("months", {})
            for month in self.months:
                status = months_data.get(month, STATUS_NOT_STUDYING)
                if status == STATUS_PAID:
                    # Используем комбинацию student_id и fio для уникальности
                    student_id = record.get("student_id", "")
                    fio = record.get("fio", "")
                    key = f"{student_id}_{fio}"
                    paid_students.add(key)
                    break  # Достаточно одного месяца с оплатой

        return len(paid_students)

    def _count_duplicate_children(self, only_paid: bool = False) -> int:
        """
        Подсчитывает количество детей в семьях с 2+ детьми (для расчета скидки)
        
        Args:
            only_paid: Учитывать только оплативших
        
        Returns:
            Количество детей в семьях с 2+ детьми
        """
        phone_counter = Counter()

        # Определяем, какие ученики учитываются
        students_to_count = set()
        for record in self.merged_data:
            if only_paid:
                # Проверяем, есть ли хотя бы один месяц с оплатой
                has_paid = False
                months_data = record.get("months", {})
                for month in self.months:
                    status = months_data.get(month, STATUS_NOT_STUDYING)
                    if status == STATUS_PAID:
                        has_paid = True
                        break
                if not has_paid:
                    continue

            student_id = record.get("student_id", "")
            fio = record.get("fio", "")
            key = f"{student_id}_{fio}"
            students_to_count.add(key)

        # Подсчитываем номера телефонов
        for record in self.merged_data:
            student_id = record.get("student_id", "")
            fio = record.get("fio", "")
            key = f"{student_id}_{fio}"

            if key not in students_to_count:
                continue

            phone = record.get("phone", "").strip()
            if phone:
                phone_counter[phone] += 1

        # Считаем общее количество детей в семьях с 2+ детьми
        total_children = 0
        for phone, count in phone_counter.items():
            if count >= MIN_CHILDREN_FOR_DISCOUNT:
                total_children += count

        return total_children

    def calculate_finances(self, month: Optional[str] = None) -> None:
        """Рассчитывает финансовые показатели за один месяц (по умолчанию — текущий месяц)."""
        if month is None:
            month = self._get_current_month()

        # считаем по месячным totals (ниже в export_excel мы делаем то же самое)
        totals_for_month = {
            "paid": 0,
            "not_paid": 0,
            "deferred": 0,
            "wrote": 0
        }

        for record in self.merged_data:
            status = record.get("months", {}).get(month, STATUS_NOT_STUDYING)
            if status == STATUS_PAID:
                totals_for_month["paid"] += 1
            elif status == STATUS_NOT_PAID:
                totals_for_month["not_paid"] += 1
            elif status == STATUS_DEFERRED:
                totals_for_month["deferred"] += 1
            elif status == STATUS_WROTE:
                totals_for_month["wrote"] += 1

        # активные в этом месяце (кроме "Не учился")
        total_students_month = sum(totals_for_month.values())
        paid_students_month = totals_for_month["paid"]

        # дубли по телефону
        total_duplicate_children = self._count_duplicate_children_for_month(month, only_paid=False)
        paid_duplicate_children = self._count_duplicate_children_for_month(month, only_paid=True)

        self.finances["expected_turnover"] = (
                total_students_month * PRICE_PER_MONTH
                - DISCOUNT_PER_CHILD * total_duplicate_children
        )
        self.finances["current_turnover"] = (
                paid_students_month * PRICE_PER_MONTH
                - DISCOUNT_PER_CHILD * paid_duplicate_children
        )
        self.finances["debt"] = (
                self.finances["expected_turnover"] - self.finances["current_turnover"]
        )

    def export_excel(self) -> Path:
        """Экспортирует данные в Excel файл"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет по оплатам"

        # Стили
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        total_font = Font(bold=True)

        # === Цвета статусов ===
        status_colors = {
            STATUS_PAID: "92D050",  # зеленый
            STATUS_NOT_PAID: "FF0000",  # красный
            STATUS_DEFERRED: "FFD966",  # желтый
            STATUS_WROTE: "5B9BD5",  # синий
            STATUS_NOT_STUDYING: "C0C0C0"  # серый
        }

        # Цвета финансов
        finance_colors = {
            "expected_turnover": "5B9BD5",  # синий
            "current_turnover": "92D050",  # зеленый
            "debt": "FF0000"  # красный
        }

        # Заголовки основной таблицы
        headers = ["ФИО", "Телефон", "URL профиля", "URL оплаты"]
        headers.extend(self.months)
        headers.append("Комментарий")

        # Записываем заголовки
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # === Записываем учеников ===
        row_idx = 2
        for record in self.merged_data:
            ws.cell(row=row_idx, column=1, value=record.get("fio", ""))
            ws.cell(row=row_idx, column=2, value=record.get("phone", ""))
            ws.cell(row=row_idx, column=3, value=record.get("student_url", ""))
            ws.cell(row=row_idx, column=4, value=record.get("payment_url", ""))

            # Месяцы
            months_data = record.get("months", {})
            for month_idx, month in enumerate(self.months, 5):
                status = months_data.get(month, STATUS_NOT_STUDYING)
                cell = ws.cell(row=row_idx, column=month_idx, value=status)

                # Окрашивание статуса
                fill_color = status_colors.get(status)
                if fill_color:
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            # Комментарий
            comment_col = 5 + len(self.months)
            ws.cell(row=row_idx, column=comment_col, value=record.get("comment", ""))

            row_idx += 1

        # === Подсчёт итогов по месяцам ===
        total_row = row_idx

        month_totals = {
            month: {"paid": 0, "not_paid": 0, "deferred": 0, "wrote": 0}
            for month in self.months
        }

        for record in self.merged_data:
            for month in self.months:
                status = record["months"][month]
                if status == STATUS_PAID:
                    month_totals[month]["paid"] += 1
                elif status == STATUS_NOT_PAID:
                    month_totals[month]["not_paid"] += 1
                elif status == STATUS_DEFERRED:
                    month_totals[month]["deferred"] += 1
                elif status == STATUS_WROTE:
                    month_totals[month]["wrote"] += 1

        # === Нижний блок — суммарно по всем месяцам ===
        summary_start = total_row + 2

        metrics = [
            ("Всего учеников", lambda t: t["paid"] + t["not_paid"] + t["deferred"] + t["wrote"]),
            ("Оплатили", lambda t: t["paid"]),
            ("Не оплатили", lambda t: t["not_paid"]),
            ("Отсрочка", lambda t: t["deferred"]),
            ("Написали", lambda t: t["wrote"]),
        ]

        # Сумма за все месяцы
        totals_all = {
            "total_students": sum(metrics[0][1](month_totals[m]) for m in self.months),
            "paid": sum(month_totals[m]["paid"] for m in self.months),
            "not_paid": sum(month_totals[m]["not_paid"] for m in self.months),
            "deferred": sum(month_totals[m]["deferred"] for m in self.months),
            "wrote": sum(month_totals[m]["wrote"] for m in self.months),
        }

        # Финансы по месяцам и суммарно
        finance_all = {"expected_turnover": 0, "current_turnover": 0, "debt": 0}
        finance_by_month = {}

        for month in self.months:
            self.calculate_finances(month)
            finance_by_month[month] = dict(self.finances)
            finance_all["expected_turnover"] += self.finances["expected_turnover"]
            finance_all["current_turnover"] += self.finances["current_turnover"]
            finance_all["debt"] += self.finances["debt"]

        # === Формируем строки метрик (A, B, E..N) ===
        current_row = summary_start

        for name, getter in metrics:
            ws.cell(row=current_row, column=1, value=name).font = total_font

            # Сумма всех месяцев (столбец B)
            key = {
                "Всего учеников": "total_students",
                "Оплатили": "paid",
                "Не оплатили": "not_paid",
                "Отсрочка": "deferred",
                "Написали": "wrote"
            }[name]

            ws.cell(row=current_row, column=2, value=totals_all[key])

            # C и D пустые
            ws.cell(row=current_row, column=3, value="")
            ws.cell(row=current_row, column=4, value="")

            # Значения по каждому месяцу
            for idx_m, month in enumerate(self.months, start=5):
                ws.cell(row=current_row, column=idx_m, value=getter(month_totals[month]))

            current_row += 1

        # === Финансы ===

        # Ожидаемый оборот
        ws.cell(row=current_row, column=1, value="Ожидаемый оборот").font = total_font
        cell = ws.cell(row=current_row, column=2, value=finance_all["expected_turnover"])
        cell.fill = PatternFill(start_color=finance_colors["expected_turnover"],
                                end_color=finance_colors["expected_turnover"], fill_type="solid")

        for idx_m, month in enumerate(self.months, start=5):
            c = ws.cell(row=current_row, column=idx_m, value=finance_by_month[month]["expected_turnover"])
            c.fill = PatternFill(start_color=finance_colors["expected_turnover"],
                                 end_color=finance_colors["expected_turnover"], fill_type="solid")

        current_row += 1

        # Оборот сейчас
        ws.cell(row=current_row, column=1, value="Оборот сейчас").font = total_font
        cell = ws.cell(row=current_row, column=2, value=finance_all["current_turnover"])
        cell.fill = PatternFill(start_color=finance_colors["current_turnover"],
                                end_color=finance_colors["current_turnover"], fill_type="solid")

        for idx_m, month in enumerate(self.months, start=5):
            c = ws.cell(row=current_row, column=idx_m, value=finance_by_month[month]["current_turnover"])
            c.fill = PatternFill(start_color=finance_colors["current_turnover"],
                                 end_color=finance_colors["current_turnover"], fill_type="solid")

        current_row += 1

        # Долг
        ws.cell(row=current_row, column=1, value="Долг").font = total_font
        cell = ws.cell(row=current_row, column=2, value=finance_all["debt"])
        cell.fill = PatternFill(start_color=finance_colors["debt"], end_color=finance_colors["debt"], fill_type="solid")

        for idx_m, month in enumerate(self.months, start=5):
            c = ws.cell(row=current_row, column=idx_m, value=finance_by_month[month]["debt"])
            c.fill = PatternFill(start_color=finance_colors["debt"], end_color=finance_colors["debt"],
                                 fill_type="solid")

        # Ширина столбцов
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 50

        for idx in range(5, 5 + len(self.months)):
            ws.column_dimensions[get_column_letter(idx)].width = 20

        ws.column_dimensions[get_column_letter(5 + len(self.months))].width = 30

        output_path = self.base_path / "payments_report.xlsx"
        wb.save(output_path)
        return output_path

    def build_summary_text(self) -> str:
        """Формирует текстовый отчет для Telegram (по текущему месяцу)"""
        current_month = self._get_current_month()
        # Всего учеников в текущем месяце (кроме "Не учился")
        total_students = (
                self.stats["paid"] +
                self.stats["not_paid"] +
                self.stats["deferred"] +
                self.stats["wrote"]
        )

        lines = [
            self.city_name,
            f"Месяц: {current_month}",
            "",
            f"Всего учеников: {total_students}",
            "",
            f"Оплатили: {self.stats['paid']}",
            f"Написали: {self.stats['wrote']}",
            f"Не оплатили: {self.stats['not_paid']}",
            f"Отсрочка: {self.stats['deferred']}",
            "",
            f"Ожидаемый оборот: {self.finances['expected_turnover']:,}".replace(",", " "),
            f"Оборот сейчас: {self.finances['current_turnover']:,}".replace(",", " "),
            f"Долг: {self.finances['debt']:,}".replace(",", " ")
        ]

        return "\n".join(lines)

    def generate_report(self) -> Tuple[str, Path]:
        """
        Генерирует полный отчет
        
        Returns:
            Tuple[str, Path]: (текстовый отчет, путь к Excel файлу)
        """
        self.load_data()
        self.merge_data()
        self.calculate_stats()
        self.calculate_finances()

        summary_text = self.build_summary_text()
        excel_path = self.export_excel()

        return summary_text, excel_path


def generate_payments_report(city_name: str) -> Tuple[str, Path]:
    """
    Генерирует отчет по оплатам для города
    
    Args:
        city_name: Название города (русское) или "all" для всех городов
    
    Returns:
        Tuple[str, Path]: (текстовый отчет, путь к Excel файлу)
    """
    if city_name == "all":
        return generate_all_cities_payments_report()
    
    generator = PaymentsReportGenerator(city_name)
    return generator.generate_report()


def generate_all_cities_payments_report() -> Tuple[str, Path]:
    """
    Генерирует общий отчет по оплатам для всех городов
    Использует те же правила, что и простой отчет, но с добавлением столбца "Город"
    
    Returns:
        Tuple[str, Path]: (текстовый отчет, путь к Excel файлу)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    from collections import Counter
    
    # Собираем данные по всем городам
    all_cities_data = []
    all_months = set()
    
    # Собираем данные из всех городов
    for city_name in CITIES:
        try:
            generator = PaymentsReportGenerator(city_name)
            generator.load_data()
            generator.merge_data()
            
            # Собираем месяцы
            all_months.update(generator.months)
            
            # Добавляем данные города с указанием города
            for record in generator.merged_data:
                record_with_city = dict(record)
                record_with_city["city"] = city_name
                all_cities_data.append(record_with_city)
        except Exception as e:
            print(f"Ошибка при загрузке данных для города {city_name}: {e}")
            continue
    
    # Сортируем месяцы
    month_order = [
        "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
    ]
    sorted_months = [m for m in month_order if m in all_months]
    
    # Используем те же методы расчета, что и в простом отчете
    # Определяем текущий месяц (используем метод из класса PaymentsReportGenerator)
    def _get_current_month() -> str:
        """Возвращает название текущего месяца, который есть в sorted_months."""
        now_month_index = datetime.now().month - 1  # 0..11
        current_name = month_order[now_month_index]
        
        if current_name in sorted_months:
            return current_name
        
        # Если текущего месяца нет в данных, берем последний доступный месяц
        if sorted_months:
            return sorted_months[-1]
        
        return ""
    
    # Проверяем, что есть данные
    if not sorted_months:
        raise ValueError("Нет данных о месяцах для генерации отчета")
    
    if not all_cities_data:
        raise ValueError("Нет данных об учениках для генерации отчета")
    
    current_month = _get_current_month()
    
    if not current_month:
        # Если не удалось определить текущий месяц, берем первый доступный
        current_month = sorted_months[0] if sorted_months else ""
        if not current_month:
            raise ValueError("Не удалось определить месяц для отчета")
    
    stats = {
        "paid": 0,
        "wrote": 0,
        "not_paid": 0,
        "deferred": 0
    }
    
    for record in all_cities_data:
        months_data = record.get("months", {})
        if not months_data:
            continue
        status = months_data.get(current_month, STATUS_NOT_STUDYING)
        
        if status == STATUS_PAID:
            stats["paid"] += 1
        elif status == STATUS_WROTE:
            stats["wrote"] += 1
        elif status == STATUS_NOT_PAID:
            stats["not_paid"] += 1
        elif status == STATUS_DEFERRED:
            stats["deferred"] += 1
    
    # Рассчитываем финансы по текущему месяцу (те же правила)
    total_students_month = stats["paid"] + stats["not_paid"] + stats["deferred"] + stats["wrote"]
    
    # Подсчет дублей по телефону для текущего месяца
    phone_counter = Counter()
    for record in all_cities_data:
        months_data = record.get("months", {})
        status = months_data.get(current_month, STATUS_NOT_STUDYING)
        
        if status == STATUS_NOT_STUDYING:
            continue
        
        phone = record.get("phone", "").strip()
        if phone:
            phone_counter[phone] += 1
    
    total_duplicate_children = sum(count for phone, count in phone_counter.items() if count >= MIN_CHILDREN_FOR_DISCOUNT)
    
    # Подсчет дублей только для оплативших
    paid_phone_counter = Counter()
    for record in all_cities_data:
        months_data = record.get("months", {})
        status = months_data.get(current_month, STATUS_NOT_STUDYING)
        
        if status != STATUS_PAID:
            continue
        
        phone = record.get("phone", "").strip()
        if phone:
            paid_phone_counter[phone] += 1
    
    paid_duplicate_children = sum(count for phone, count in paid_phone_counter.items() if count >= MIN_CHILDREN_FOR_DISCOUNT)
    
    # Расчет финансов (те же формулы)
    expected_turnover = (
        total_students_month * PRICE_PER_MONTH
        - DISCOUNT_PER_CHILD * total_duplicate_children
    )
    current_turnover = (
        stats["paid"] * PRICE_PER_MONTH
        - DISCOUNT_PER_CHILD * paid_duplicate_children
    )
    debt = expected_turnover - current_turnover
    
    # Формируем текстовый отчет (те же правила)
    total_students = stats["paid"] + stats["not_paid"] + stats["deferred"] + stats["wrote"]
    
    summary_lines = [
        "Общий отчёт по оплатам (все города)",
        f"Месяц: {current_month}",
        "",
        f"Всего учеников: {total_students}",
        "",
        f"Оплатили: {stats['paid']}",
        f"Написали: {stats['wrote']}",
        f"Не оплатили: {stats['not_paid']}",
        f"Отсрочка: {stats['deferred']}",
        "",
        f"Ожидаемый оборот: {expected_turnover:,}".replace(",", " "),
        f"Оборот сейчас: {current_turnover:,}".replace(",", " "),
        f"Долг: {debt:,}".replace(",", " ")
    ]
    
    summary_text = "\n".join(summary_lines)
    
    # Создаем Excel файл (те же стили и структура, но с добавлением столбца "Город")
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет по оплатам"
    
    # Стили (те же, что и в простом отчете)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    total_font = Font(bold=True)
    
    # Цвета статусов (те же)
    status_colors = {
        STATUS_PAID: "92D050",  # зеленый
        STATUS_NOT_PAID: "FF0000",  # красный
        STATUS_DEFERRED: "FFD966",  # желтый
        STATUS_WROTE: "5B9BD5",  # синий
        STATUS_NOT_STUDYING: "C0C0C0"  # серый
    }
    
    # Цвета финансов (те же)
    finance_colors = {
        "expected_turnover": "5B9BD5",  # синий
        "current_turnover": "92D050",  # зеленый
        "debt": "FF0000"  # красный
    }
    
    # Заголовки (добавляем "Город" перед ФИО)
    headers = ["Город", "ФИО", "Телефон", "URL профиля", "URL оплаты"]
    headers.extend(sorted_months)
    headers.append("Комментарий")
    
    # Записываем заголовки
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Записываем учеников
    row_idx = 2
    for record in all_cities_data:
        ws.cell(row=row_idx, column=1, value=record.get("city", ""))  # Город
        ws.cell(row=row_idx, column=2, value=record.get("fio", ""))  # ФИО
        ws.cell(row=row_idx, column=3, value=record.get("phone", ""))  # Телефон
        ws.cell(row=row_idx, column=4, value=record.get("student_url", ""))  # URL профиля
        ws.cell(row=row_idx, column=5, value=record.get("payment_url", ""))  # URL оплаты
        
        # Месяцы (начинаем с колонки 6, так как добавили "Город")
        months_data = record.get("months", {})
        for month_idx, month in enumerate(sorted_months, 6):
            status = months_data.get(month, STATUS_NOT_STUDYING)
            cell = ws.cell(row=row_idx, column=month_idx, value=status)
            
            # Окрашивание статуса
            fill_color = status_colors.get(status)
            if fill_color:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # Комментарий
        comment_col = 6 + len(sorted_months)
        ws.cell(row=row_idx, column=comment_col, value=record.get("comment", ""))
        
        row_idx += 1
    
    # Подсчет итогов по месяцам (те же правила)
    total_row = row_idx
    
    month_totals = {
        month: {"paid": 0, "not_paid": 0, "deferred": 0, "wrote": 0}
        for month in sorted_months
    }
    
    for record in all_cities_data:
        months_data = record.get("months", {})
        if not months_data:
            continue
        for month in sorted_months:
            status = months_data.get(month, STATUS_NOT_STUDYING)
            if status == STATUS_PAID:
                month_totals[month]["paid"] += 1
            elif status == STATUS_NOT_PAID:
                month_totals[month]["not_paid"] += 1
            elif status == STATUS_DEFERRED:
                month_totals[month]["deferred"] += 1
            elif status == STATUS_WROTE:
                month_totals[month]["wrote"] += 1
    
    # Нижний блок — суммарно по всем месяцам (те же метрики)
    summary_start = total_row + 2
    
    metrics = [
        ("Всего учеников", lambda t: t["paid"] + t["not_paid"] + t["deferred"] + t["wrote"]),
        ("Оплатили", lambda t: t["paid"]),
        ("Не оплатили", lambda t: t["not_paid"]),
        ("Отсрочка", lambda t: t["deferred"]),
        ("Написали", lambda t: t["wrote"]),
    ]
    
    # Сумма за все месяцы
    totals_all = {
        "total_students": sum(metrics[0][1](month_totals[m]) for m in sorted_months),
        "paid": sum(month_totals[m]["paid"] for m in sorted_months),
        "not_paid": sum(month_totals[m]["not_paid"] for m in sorted_months),
        "deferred": sum(month_totals[m]["deferred"] for m in sorted_months),
        "wrote": sum(month_totals[m]["wrote"] for m in sorted_months),
    }
    
    # Финансы по месяцам и суммарно (те же расчеты)
    finance_all = {"expected_turnover": 0, "current_turnover": 0, "debt": 0}
    finance_by_month = {}
    
    for month in sorted_months:
        # Расчет финансов для каждого месяца
        month_total = month_totals[month]["paid"] + month_totals[month]["not_paid"] + month_totals[month]["deferred"] + month_totals[month]["wrote"]
        
        # Подсчет дублей для этого месяца
        month_phone_counter = Counter()
        month_paid_phone_counter = Counter()
        
        for record in all_cities_data:
            months_data = record.get("months", {})
            status = months_data.get(month, STATUS_NOT_STUDYING)
            
            if status != STATUS_NOT_STUDYING:
                phone = record.get("phone", "").strip()
                if phone:
                    month_phone_counter[phone] += 1
                    if status == STATUS_PAID:
                        month_paid_phone_counter[phone] += 1
        
        month_duplicate = sum(count for phone, count in month_phone_counter.items() if count >= MIN_CHILDREN_FOR_DISCOUNT)
        month_paid_duplicate = sum(count for phone, count in month_paid_phone_counter.items() if count >= MIN_CHILDREN_FOR_DISCOUNT)
        
        month_expected = month_total * PRICE_PER_MONTH - DISCOUNT_PER_CHILD * month_duplicate
        month_current = month_totals[month]["paid"] * PRICE_PER_MONTH - DISCOUNT_PER_CHILD * month_paid_duplicate
        month_debt = month_expected - month_current
        
        finance_by_month[month] = {
            "expected_turnover": month_expected,
            "current_turnover": month_current,
            "debt": month_debt
        }
        
        finance_all["expected_turnover"] += month_expected
        finance_all["current_turnover"] += month_current
        finance_all["debt"] += month_debt
    
    # Формируем строки метрик (те же, что и в простом отчете)
    current_row = summary_start
    
    for name, getter in metrics:
        ws.cell(row=current_row, column=1, value=name).font = total_font
        
        # Сумма всех месяцев (столбец B)
        key = {
            "Всего учеников": "total_students",
            "Оплатили": "paid",
            "Не оплатили": "not_paid",
            "Отсрочка": "deferred",
            "Написали": "wrote"
        }[name]
        ws.cell(row=current_row, column=2, value=totals_all[key])
        
        # C, D, E пустые (Город, ФИО, Телефон)
        ws.cell(row=current_row, column=3, value="")
        ws.cell(row=current_row, column=4, value="")
        ws.cell(row=current_row, column=5, value="")
        
        # Значения по каждому месяцу (начинаем с колонки 6)
        for idx_m, month in enumerate(sorted_months, start=6):
            ws.cell(row=current_row, column=idx_m, value=getter(month_totals[month]))
        
        current_row += 1
    
    # Финансы (те же строки)
    # Ожидаемый оборот
    ws.cell(row=current_row, column=1, value="Ожидаемый оборот").font = total_font
    cell = ws.cell(row=current_row, column=2, value=finance_all["expected_turnover"])
    cell.fill = PatternFill(start_color=finance_colors["expected_turnover"],
                            end_color=finance_colors["expected_turnover"], fill_type="solid")
    
    ws.cell(row=current_row, column=3, value="")
    ws.cell(row=current_row, column=4, value="")
    ws.cell(row=current_row, column=5, value="")
    
    for idx_m, month in enumerate(sorted_months, start=6):
        c = ws.cell(row=current_row, column=idx_m, value=finance_by_month[month]["expected_turnover"])
        c.fill = PatternFill(start_color=finance_colors["expected_turnover"],
                             end_color=finance_colors["expected_turnover"], fill_type="solid")
    
    current_row += 1
    
    # Оборот сейчас
    ws.cell(row=current_row, column=1, value="Оборот сейчас").font = total_font
    cell = ws.cell(row=current_row, column=2, value=finance_all["current_turnover"])
    cell.fill = PatternFill(start_color=finance_colors["current_turnover"],
                            end_color=finance_colors["current_turnover"], fill_type="solid")
    
    ws.cell(row=current_row, column=3, value="")
    ws.cell(row=current_row, column=4, value="")
    ws.cell(row=current_row, column=5, value="")
    
    for idx_m, month in enumerate(sorted_months, start=6):
        c = ws.cell(row=current_row, column=idx_m, value=finance_by_month[month]["current_turnover"])
        c.fill = PatternFill(start_color=finance_colors["current_turnover"],
                             end_color=finance_colors["current_turnover"], fill_type="solid")
    
    current_row += 1
    
    # Долг
    ws.cell(row=current_row, column=1, value="Долг").font = total_font
    cell = ws.cell(row=current_row, column=2, value=finance_all["debt"])
    cell.fill = PatternFill(start_color=finance_colors["debt"], end_color=finance_colors["debt"], fill_type="solid")
    
    ws.cell(row=current_row, column=3, value="")
    ws.cell(row=current_row, column=4, value="")
    ws.cell(row=current_row, column=5, value="")
    
    for idx_m, month in enumerate(sorted_months, start=6):
        c = ws.cell(row=current_row, column=idx_m, value=finance_by_month[month]["debt"])
        c.fill = PatternFill(start_color=finance_colors["debt"], end_color=finance_colors["debt"],
                             fill_type="solid")
    
    # Ширина столбцов (те же, но с учетом добавленного столбца "Город")
    ws.column_dimensions["A"].width = 15  # Город
    ws.column_dimensions["B"].width = 30  # ФИО
    ws.column_dimensions["C"].width = 18  # Телефон
    ws.column_dimensions["D"].width = 50  # URL профиля
    ws.column_dimensions["E"].width = 50  # URL оплаты
    
    for idx in range(6, 6 + len(sorted_months)):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    ws.column_dimensions[get_column_letter(6 + len(sorted_months))].width = 30  # Комментарий
    
    # Сохраняем файл
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"payments_report_all_cities_{timestamp}.xlsx"
    excel_path = ROOT_DIR / "temp" / filename
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)
    
    return summary_text, excel_path
